# with open ("Python\\Lesson_10\\file_1.txt", "w", encoding="utf_8") as file_1:
#     file_1.write("Snow on the ground.\nSnow on the tree.\nSnow on the house.\nSnow on me!")

# with open ("Python\\Lesson_10\\file_1.txt", "r", encoding="utf_8") as file_1:
#     data = file_1.read()

# with open("Python\\Lesson_10\\file_2.txt", "w", encoding="utf_8") as file_2:
#     file_2.write(data)

import openpyxl as xl
import json


# new = xl.Workbook()
# new_active = new.active

# for i in range(1, 11):
#     for j in range(10):
#         new_active.cell(row=j+1, column=i, value=i * (j+1))

# new.save('new_file.xlsx')
# new.close()

with open("db.json") as db:
    data = json.load(db)

new = xl.Workbook()
new_active = new.active

new_row = 0
new_column = 1

for i in data['movies']:
    new_row += 1
    if new_row == 1:
        new_active.cell(row=new_row, column=new_column, value= 'Tile')
        new_active.cell(row=new_row, column=new_column + 1, value= 'Year')
        new_active.cell(row=new_row, column=new_column + 2, value= 'Genres')
        new_active.cell(row=new_row, column=new_column + 3, value= 'Director')
    else:
        new_active.cell(row=new_row, column=new_column, value= i['title'])
        new_active.cell(row=new_row, column=new_column + 1, value= i['year'])
        new_active.cell(row=new_row, column=new_column + 2, value= ', '.join(i['genres']))
        new_active.cell(row=new_row, column=new_column + 3, value= i['director'])


new.save('new_file.xlsx')
new.close()